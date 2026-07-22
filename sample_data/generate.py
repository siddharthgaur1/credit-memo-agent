"""Generate three synthetic borrower files as digital PDFs + Excel.

EVERYTHING HERE IS INVENTED. See README.md in this directory.

Figures are held in lakhs (the convention Indian SME statements are printed in)
and written that way into the documents; the extractor converts to absolute
rupees. Layouts mimic Schedule III presentation closely enough to exercise the
table extraction, including a workbook whose figures start below a title block.
"""

from __future__ import annotations

import calendar
from datetime import date
from pathlib import Path

OUT = Path(__file__).resolve().parent / "out"

# Dates are relative to today so the demo file never rots into "everything is
# stale". The Indian financial year ends 31 March.
TODAY = date.today()
LAST_FY_END_YEAR = TODAY.year if TODAY.month >= 4 else TODAY.year - 1


def fy(offset: int) -> str:
    """FY label `offset` years before the last completed financial year."""
    return f"FY{(LAST_FY_END_YEAR - offset) % 100:02d}"


def month_end(months_ago: int) -> date:
    """End of the month `months_ago` before the current one."""
    total = TODAY.year * 12 + (TODAY.month - 1) - months_ago
    year, month = divmod(total, 12)
    return date(year, month + 1, calendar.monthrange(year, month + 1)[1])


def month_label(d: date) -> str:
    return f"{calendar.month_abbr[d.month]} {d.year}"


FY_END = {fy(i): f"31 March {LAST_FY_END_YEAR - i}" for i in range(0, 8)}
RECENT_MONTHS = [month_label(month_end(i)) for i in (4, 3, 2, 1)]


def bank_period(years_ago: int = 0) -> str:
    end = month_end(1 + years_ago * 12)
    start = month_end(12 + years_ago * 12)
    return f"{month_label(start)} - {month_label(end)}"

# --------------------------------------------------------------------------- #
# Borrower data (Rs. in lakhs)
# --------------------------------------------------------------------------- #

HEALTHY = {
    "slug": "suryodaya_engineering",
    "name": "Suryodaya Engineering Works Pvt Ltd",
    "pan": "AABCS1234K",
    "gstin": "27AABCS1234K1ZR",
    "address": "Plot 14, MIDC Phase II, Ahmednagar, Maharashtra 414111",
    "constitution": "Private Limited Company",
    "facility": "Term loan of Rs. 180.00 lakhs for purchase of a CNC turning centre",
    "years": [fy(1), fy(0)],
    "bs": {
        "Share capital": [100, 100],
        "Reserves and surplus": [330, 420],
        "Net worth": [430, 520],
        "Long-term borrowings": [300, 260],
        "Other non-current liabilities": [35, 40],
        "Short-term borrowings": [160, 180],
        "Trade payables": [170, 190],
        "Other current liabilities": [50, 60],
        "Total current liabilities": [380, 430],
        "Net fixed assets": [545, 520],
        "Other non-current assets": [25, 30],
        "Inventories": [240, 260],
        "Trade receivables": [265, 300],
        "Cash and bank balances": [55, 90],
        "Other current assets": [15, 50],
        "Total current assets": [575, 700],
        "Total assets": [1145, 1250],
    },
    "pl": {
        "Revenue from operations": [2050, 2400],
        "Cost of goods sold": [1435, 1656],
        "Gross profit": [615, 744],
        "Employee benefit expense": [215, 240],
        "Other expenses": [215, 264],
        "EBITDA": [185, 240],
        "Other income": [8, 12],
        "Depreciation and amortisation": [70, 75],
        "Finance costs": [66, 62],
        "Profit before tax": [57, 115],
        "Tax expense": [14, 29],
        "Profit after tax": [43, 86],
    },
    "debt_service": {"principal_repayment": [58, 55], "proposed_emi_annual": [0, 0]},
    "bank": {
        "period": bank_period(),
        "avg_monthly_balance": 41.5,
        "minimum_balance": 6.2,
        "total_credits": 2515.0,
        "total_debits": 2478.0,
        "inward_returns": 0,
        "outward_returns": 1,
        "sanctioned_limit": 200.0,
        "peak_utilisation": 168.0,
    },
    "docs": ["gst", "itr", "kyc", "project", "sanction"],
    "gst_months": RECENT_MONTHS,
    "stale": False,
}

STRESSED = {
    "slug": "vaishnavi_polymers",
    "name": "Vaishnavi Polymers LLP",
    "pan": "AAEFV5678M",
    "gstin": "29AAEFV5678M1Z8",
    "address": "Shed 7, Peenya Industrial Area, Bengaluru, Karnataka 560058",
    "constitution": "Limited Liability Partnership",
    "facility": "Term loan of Rs. 120.00 lakhs for a blow-moulding line",
    "years": [fy(1), fy(0)],
    "bs": {
        "Share capital": [60, 60],
        "Reserves and surplus": [95, 40],
        "Net worth": [155, 100],
        "Long-term borrowings": [380, 340],
        "Other non-current liabilities": [18, 20],
        "Short-term borrowings": [260, 300],
        "Trade payables": [240, 280],
        "Other current liabilities": [62, 70],
        "Total current liabilities": [562, 650],
        "Net fixed assets": [470, 430],
        "Other non-current assets": [18, 20],
        "Inventories": [300, 330],
        "Trade receivables": [285, 300],
        "Cash and bank balances": [22, 10],
        "Other current assets": [20, 20],
        "Total current assets": [627, 660],
        "Total assets": [1115, 1110],
    },
    "pl": {
        "Revenue from operations": [1800, 1420],
        "Cost of goods sold": [1400, 1150],
        "Gross profit": [400, 270],
        "Employee benefit expense": [140, 130],
        "Other expenses": [130, 108],
        "EBITDA": [130, 32],
        "Other income": [5, 3],
        "Depreciation and amortisation": [66, 62],
        "Finance costs": [82, 88],
        "Profit before tax": [-13, -115],
        "Tax expense": [0, 0],
        "Profit after tax": [-13, -115],
    },
    "debt_service": {"principal_repayment": [42, 45], "proposed_emi_annual": [0, 0]},
    "bank": {
        "period": bank_period(),
        "avg_monthly_balance": 3.8,
        "minimum_balance": -12.4,
        "total_credits": 1462.0,
        "total_debits": 1481.0,
        "inward_returns": 4,
        "outward_returns": 6,
        "sanctioned_limit": 300.0,
        "peak_utilisation": 295.0,
    },
    "docs": ["gst", "itr", "kyc", "project", "sanction"],
    "gst_months": RECENT_MONTHS,
    "stale": False,
}

# Deliberately broken: FY21 total assets is printed as 860 while the components
# sum to 830 and equity + liabilities is 830. The tool must FLAG this, not fix it.
INCOMPLETE = {
    "slug": "meenakshi_traders",
    "name": "Meenakshi Traders",
    "pan": "AFHPM9012Q",
    "gstin": "33AFHPM9012Q1ZK",
    "address": "23 Nethaji Road, Coimbatore, Tamil Nadu 641001",
    "constitution": "Proprietorship",
    "facility": "Term loan of Rs. 60.00 lakhs for godown expansion",
    "years": [fy(5), fy(4)],
    "bs": {
        "Share capital": [50, 50],
        "Reserves and surplus": [185, 210],
        "Net worth": [235, 260],
        "Long-term borrowings": [215, 200],
        "Other non-current liabilities": [12, 15],
        "Short-term borrowings": [140, 150],
        "Trade payables": [145, 160],
        "Other current liabilities": [40, 45],
        "Total current liabilities": [325, 355],
        "Net fixed assets": [365, 380],
        "Other non-current assets": [18, 20],
        "Inventories": [170, 180],
        "Trade receivables": [180, 190],
        "Cash and bank balances": [32, 40],
        "Other current assets": [12, 20],
        "Total current assets": [394, 430],
        "Total assets": [787, 860],  # FY21 does not balance: components sum to 830
    },
    # Depreciation is genuinely absent from FY21 -- DSCR and ROCE must come back
    # as "insufficient data", never as a guess.
    "pl": {
        "Revenue from operations": [1120, 1180],
        "Cost of goods sold": [890, 940],
        "Gross profit": [230, 240],
        "Employee benefit expense": [70, 74],
        "Other expenses": [92, 96],
        "EBITDA": [68, 70],
        "Other income": [4, 4],
        "Depreciation and amortisation": [28, None],
        "Finance costs": [34, 33],
        "Profit before tax": [10, 41],
        "Tax expense": [3, 10],
        "Profit after tax": [7, 31],
    },
    "debt_service": {"principal_repayment": [30, 30], "proposed_emi_annual": [0, 0]},
    "bank": {
        "period": bank_period(4),
        "avg_monthly_balance": 12.1,
        "minimum_balance": 0.9,
        "total_credits": 1198.0,
        "total_debits": 1191.0,
        "inward_returns": 1,
        "outward_returns": 0,
        "sanctioned_limit": 150.0,
        "peak_utilisation": 138.0,
    },
    "docs": ["itr", "kyc", "project"],  # no GST returns, no sanction letter
    "gst_months": [],
    "stale": True,
}

BORROWERS = [HEALTHY, STRESSED, INCOMPLETE]



# --------------------------------------------------------------------------- #
# PDF helpers
# --------------------------------------------------------------------------- #


def _styles():
    from reportlab.lib.styles import getSampleStyleSheet

    return getSampleStyleSheet()


def _table(rows: list[list[str]], col_widths=None):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    t = Table(rows, colWidths=col_widths, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return t


def _pdf(path: Path, title: str, blocks: list) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate

    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(path), pagesize=A4, title=title)
    doc.build(blocks)


def _amount(v) -> str:
    if v is None:
        return ""
    return f"({abs(v):,.2f})" if v < 0 else f"{v:,.2f}"


def _para(text: str, style="Normal"):
    from reportlab.platypus import Paragraph

    return Paragraph(text, _styles()[style])


def _spacer(h=10):
    from reportlab.platypus import Spacer

    return Spacer(1, h)


# --------------------------------------------------------------------------- #
# Documents
# --------------------------------------------------------------------------- #

BS_ORDER = [
    ("EQUITY AND LIABILITIES", None),
    ("Share capital", "Share capital"),
    ("Reserves and surplus", "Reserves and surplus"),
    ("Net worth", "Net worth"),
    ("Long-term borrowings", "Long-term borrowings"),
    ("Other non-current liabilities", "Other non-current liabilities"),
    ("Short-term borrowings", "Short-term borrowings"),
    ("Trade payables", "Trade payables"),
    ("Other current liabilities", "Other current liabilities"),
    ("Total current liabilities", "Total current liabilities"),
    ("ASSETS", None),
    ("Net fixed assets", "Net fixed assets"),
    ("Other non-current assets", "Other non-current assets"),
    ("Inventories", "Inventories"),
    ("Trade receivables", "Trade receivables"),
    ("Cash and bank balances", "Cash and bank balances"),
    ("Other current assets", "Other current assets"),
    ("Total current assets", "Total current assets"),
    ("Total assets", "Total assets"),
]

PL_ORDER = [
    "Revenue from operations",
    "Cost of goods sold",
    "Gross profit",
    "Employee benefit expense",
    "Other expenses",
    "EBITDA",
    "Other income",
    "Depreciation and amortisation",
    "Finance costs",
    "Profit before tax",
    "Tax expense",
    "Profit after tax",
]


def balance_sheet_pdf(b: dict, i: int, out: Path) -> None:
    fy = b["years"][i]
    rows = [["Particulars", f"As at {FY_END[fy]}"]]
    for label, key in BS_ORDER:
        rows.append([label, "" if key is None else _amount(b["bs"][key][i])])
    _pdf(
        out / f"balance_sheet_{fy}.pdf",
        f"{b['name']} Balance Sheet {fy}",
        [
            _para(f"<b>{b['name']}</b>", "Title"),
            _para(b["address"]),
            _para(f"PAN: {b['pan']} | Constitution: {b['constitution']}"),
            _spacer(),
            _para(f"<b>BALANCE SHEET AS AT {FY_END[fy].upper()}</b>", "Heading2"),
            _para("(Rs. in lakhs)"),
            _spacer(6),
            _table(rows, col_widths=[280, 140]),
            _spacer(),
            _para("Shareholders funds and total equity and liabilities as reported above."),
            _para("Audited. For M/s Ramanathan &amp; Co., Chartered Accountants. FRN 004512S."),
        ],
    )


def pl_pdf(b: dict, i: int, out: Path) -> None:
    fy = b["years"][i]
    rows = [["Particulars", f"Year ended {FY_END[fy]}"]]
    for label in PL_ORDER:
        rows.append([label, _amount(b["pl"][label][i])])
    _pdf(
        out / f"profit_and_loss_{fy}.pdf",
        f"{b['name']} P&L {fy}",
        [
            _para(f"<b>{b['name']}</b>", "Title"),
            _para(b["address"]),
            _spacer(),
            _para(f"<b>STATEMENT OF PROFIT AND LOSS FOR THE YEAR ENDED {FY_END[fy].upper()}</b>", "Heading2"),
            _para("(Rs. in lakhs)"),
            _spacer(6),
            _table(rows, col_widths=[280, 140]),
            _spacer(),
            _para("<b>Debt service during the year</b> (Rs. in lakhs)"),
            _spacer(6),
            _table(
                [
                    ["Particulars", f"Year ended {FY_END[fy]}"],
                    ["Principal repayment of term loans", _amount(b["debt_service"]["principal_repayment"][i])],
                ],
                col_widths=[280, 140],
            ),
            _spacer(),
            _para("Audited. For M/s Ramanathan &amp; Co., Chartered Accountants. FRN 004512S."),
        ],
    )


def gst_pdf(b: dict, month: str, out: Path) -> None:
    _pdf(
        out / f"gstr3b_{month.replace(' ', '_').lower()}.pdf",
        f"GSTR-3B {month}",
        [
            _para("<b>FORM GSTR-3B</b>", "Title"),
            _para("Goods and Services Tax - Monthly Return"),
            _spacer(),
            _table(
                [
                    ["Particulars", "Value (Rs.)"],
                    ["GSTIN", b["gstin"]],
                    ["Legal name", b["name"]],
                    ["Return period", month],
                    ["Outward supplies - taxable value", "18,42,500.00"],
                    ["Integrated tax", "1,10,550.00"],
                    ["Central tax", "1,65,825.00"],
                    ["State tax", "1,65,825.00"],
                    ["Filing status", "Filed"],
                ],
                col_widths=[240, 180],
            ),
        ],
    )


def itr_pdf(b: dict, out: Path) -> None:
    start = LAST_FY_END_YEAR - (4 if b["stale"] else 0)
    ay = f"{start}-{(start + 1) % 100:02d}"
    filed = month_end(49 if b["stale"] else 1)
    _pdf(
        out / f"itr_acknowledgement_ay{ay.replace('-', '_')}.pdf",
        f"ITR Acknowledgement AY {ay}",
        [
            _para("<b>INDIAN INCOME TAX RETURN ACKNOWLEDGEMENT</b>", "Title"),
            _spacer(),
            _table(
                [
                    ["Particulars", "Details"],
                    ["Name of the assessee", b["name"]],
                    ["PAN of the assessee", b["pan"]],
                    ["Assessment year", ay],
                    ["Acknowledgement number", "3948217650" + ay[-2:]],
                    ["Gross total income", "1,14,80,000"],
                    ["Total tax paid", "28,90,000"],
                    ["Date of filing", month_label(filed)],
                ],
                col_widths=[240, 200],
            ),
        ],
    )


def kyc_pdf(b: dict, out: Path) -> None:
    _pdf(
        out / "kyc_pan_udyam_address.pdf",
        "KYC pack",
        [
            _para("<b>KYC DOCUMENT PACK</b>", "Title"),
            _spacer(),
            _table(
                [
                    ["Document", "Reference"],
                    ["Permanent Account Number", b["pan"]],
                    ["Udyam registration", "UDYAM-MH-18-0043921"],
                    ["Constitution proof", b["constitution"]],
                    ["Proof of address", b["address"]],
                    ["Aadhaar of proprietor/director (masked)", "XXXX XXXX 4417"],
                ],
                col_widths=[260, 220],
            ),
        ],
    )


def project_pdf(b: dict, out: Path) -> None:
    _pdf(
        out / "project_report_and_quotation.pdf",
        "Project report",
        [
            _para("<b>PROJECT REPORT AND PROFORMA INVOICE</b>", "Title"),
            _para(b["name"]),
            _spacer(),
            _para(f"Facility sought: {b['facility']}"),
            _spacer(6),
            _table(
                [
                    ["Cost of project", "Amount (Rs. in lakhs)"],
                    ["Plant and machinery (as per quotation)", "168.00"],
                    ["Erection and commissioning", "9.50"],
                    ["Contingency", "4.50"],
                    ["Total cost of project", "182.00"],
                    ["Means of finance - promoter contribution", "45.50"],
                    ["Means of finance - proposed term loan", "136.50"],
                ],
                col_widths=[280, 160],
            ),
            _spacer(),
            _para(
                f"Quotation dated {month_label(month_end(49 if b['stale'] else 1))}, "
                "valid for 90 days from date of issue. Supplier: Shakti Machine Tools Pvt Ltd."
            ),
        ],
    )


def sanction_pdf(b: dict, out: Path) -> None:
    _pdf(
        out / "sanction_letter_existing_facility.pdf",
        "Sanction letter",
        [
            _para("<b>SANCTION LETTER - EXISTING FACILITY</b>", "Title"),
            _spacer(),
            _para(f"To: {b['name']}, {b['address']}"),
            _spacer(6),
            _table(
                [
                    ["Terms and conditions of sanction", "Details"],
                    ["Facility", "Cash credit"],
                    ["Sanctioned amount", f"Rs. {b['bank']['sanctioned_limit']:,.2f} lakhs"],
                    ["Rate of interest", "10.75% p.a. floating"],
                    ["Repayment schedule", "On demand; annual review"],
                    ["EMI (term loan component)", "Rs. 3.90 lakhs per month"],
                    ["Security", "Hypothecation of stock and book debts"],
                ],
                col_widths=[240, 220],
            ),
        ],
    )


def bank_statement_xlsx(b: dict, out: Path) -> None:
    """Workbook with a title block above the header row and a second sheet.

    This is the layout that trips naive parsers: the real header sits on row 7.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "STATE BANK OF INDIA - ACCOUNT STATEMENT"
    ws["A2"] = b["name"]
    ws["A3"] = f"Account number: XXXXXXXX{4417 + len(b['slug'])}  IFSC: SBIN0004512"
    ws["A4"] = f"Statement period: {b['bank']['period']}"
    ws["A5"] = "All amounts in Rs. lakhs unless stated otherwise"
    ws["A6"] = ""

    rows = [
        ["Particulars", "Value"],
        ["Average monthly balance", b["bank"]["avg_monthly_balance"]],
        ["Minimum balance during period", b["bank"]["minimum_balance"]],
        ["Total credits", b["bank"]["total_credits"]],
        ["Total debits", b["bank"]["total_debits"]],
        ["Inward cheque returns (own cheques returned unpaid)", b["bank"]["inward_returns"]],
        ["Outward cheque returns (deposited cheques returned)", b["bank"]["outward_returns"]],
        ["Sanctioned limit", b["bank"]["sanctioned_limit"]],
        ["Peak utilisation", b["bank"]["peak_utilisation"]],
        ["Months covered", 12],
    ]
    for r in rows:
        ws.append(r)

    tx = wb.create_sheet("Transactions")
    tx.append(["Value date", "Narration", "Withdrawal", "Deposit", "Closing balance"])
    running = b["bank"]["avg_monthly_balance"]
    for month in range(1, 13):
        tx.append([f"{month:02d}-04-2022", "NEFT inward - customer receipt", "", 42.5, running + 42.5])
        tx.append([f"{month:02d}-04-2022", "Supplier payment", 38.0, "", running + 4.5])
        running = round(running + 4.5, 2)

    out.mkdir(parents=True, exist_ok=True)
    wb.save(out / "bank_statement_summary.xlsx")


def generate(b: dict, root: Path = OUT) -> Path:
    out = root / b["slug"]
    out.mkdir(parents=True, exist_ok=True)
    for i in range(len(b["years"])):
        balance_sheet_pdf(b, i, out)
        pl_pdf(b, i, out)
    for month in b["gst_months"]:
        gst_pdf(b, month, out)
    if "itr" in b["docs"]:
        itr_pdf(b, out)
    if "kyc" in b["docs"]:
        kyc_pdf(b, out)
    if "project" in b["docs"]:
        project_pdf(b, out)
    if "sanction" in b["docs"]:
        sanction_pdf(b, out)
    bank_statement_xlsx(b, out)
    return out


def main() -> None:
    for b in BORROWERS:
        path = generate(b)
        print(f"{b['name']:42s} -> {path} ({len(list(path.iterdir()))} files)")
    print(f"\nGenerated on {date.today().isoformat()}. All data is synthetic.")


if __name__ == "__main__":
    main()
